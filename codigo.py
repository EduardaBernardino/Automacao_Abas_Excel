from openpyxl import load_workbook

arquivo = load_workbook("Alunos.xlsx")

#print(arquivo.sheetnames) #mostra o nome das abas que tenho na minha planilha do excel. Ex: ['Planilha1', 'Planilha2']

aba_alunos = arquivo["Planilha1"]
print(aba_alunos) #mostra o objeto da aba que estou acessando


#Valor antes
valor_b2 = aba_alunos.cell(row=1, column=1).value
print(valor_b2)

#Valor apos alterações
apos = aba_alunos.cell(row=1, column=1).value = "COD"
print(apos)


arquivo.save("Alunos.xlsx") # 4. SALVA as alterações no arquivo.
