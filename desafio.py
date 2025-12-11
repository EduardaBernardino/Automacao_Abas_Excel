#percorrer o arquivo todo e verificar se a aba de cada bairro ja existe, se nao existir criar uma nova aba

#copiar os valores daquela linha e inserir no bairro correspondente

from openpyxl import load_workbook

def criar_aba(bairro, arquivo_bairros):
    if bairro not in arquivo_bairros.sheetnames:
        arquivo_bairros.create_sheet(title=bairro)
        nova_aba = arquivo_bairros[bairro]
        nova_aba["A1"].value = "Data de Nascimento"
        nova_aba["B1"].value = "Nome"
        nova_aba["C1"].value = "Bairro"

# Função de transferência
def transferir_informacoes_aba(aba_Base_Dados, aba_destino, linha_origem):
    # Encontra a próxima linha vazia na aba de destino
    linha_destino = aba_destino.max_row + 1

    # Colunas de 1 a 3 (A, B, C)
    for coluna in range (1, 4):
        celula_origem = aba_Base_Dados.cell(row=linha_origem, column=coluna)

        # 2. PEGA A CÉLULA DE DESTINO
        celula_destino = aba_destino.cell(row=linha_destino, column=coluna)

        # 3. COPIA O VALOR
        celula_destino.value = celula_origem.value


arquivo_bairros = load_workbook("Bairros.xlsx")
aba_Base_Dados = arquivo_bairros["Base_Dados"]

ultima_linha = aba_Base_Dados.max_row #faço isso para ver qual a última linha que o código terá que percorrer
print(f"Total de linhas a percorrer na Base de Dados: {ultima_linha}")


# Loop começa na linha 2 (assumindo que a linha 1 é o cabeçalho)
for linha in range(2, ultima_linha + 1):

    bairro = aba_Base_Dados.cell(row=linha, column=3).value

    # 1. Checagem de Parada (se o bairro for vazio, interrompe o loop)
    if not bairro:
        break


    # Cria a aba para o bairro (se já não existir)
    criar_aba(bairro, arquivo_bairros)


    aba_destino = arquivo_bairros[bairro]

    # Transfere a linha atual (cópia)
    transferir_informacoes_aba(aba_Base_Dados, aba_destino, linha)


arquivo_bairros.save("Bairros_Separados.xlsx")
